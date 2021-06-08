import pandas as pd
import openpyxl
import csv
import sys
import os
from dotenv import load_dotenv
from os import system, name
from datetime import date, datetime
from openpyxl.styles import Font, Fill, PatternFill  # Connect styles for text


from timesheet_entry import timesheet_entry
from hmrc_timesheet_entry import hmrc_timesheet_entry
from resource import resource
from sow import sow


class application(object):
    def __init__(self):
        self.src = ""
        self.dest = ""
        self.master_dest = ""
        self.master_sheet = ""
        self.match_count = 0
        self.entries = []
        self.hmrc_entries = []
        self.resources = []
        self.sows = []

        self.permitted_employee = [1.875, 3.75, 5.625, 7.5]
        self.permitted_contractor = [2, 4, 6, 8]
        self.dest = 'src/temp.csv'
        self.master_dest = 'src/temp_master.csv'

        self.get_project()
        self.get_timesheet_month()

    def get_config(self):
        load_dotenv('.env')

        self.source_path = os.getenv('SOURCE_PATH')
        self.source_file = os.getenv('SOURCE_FILE')
        self.sheet = os.getenv('SHEET')

        self.src = os.path.join(self.source_path, self.source_file)
        self.master = os.path.join(
            self.source_path, "HMRC billing master.xlsx")

    def clear(self):
        # for windows
        if name == 'nt':
            _ = system('cls')
        else:
            _ = system("printf '\33c\e[3J'")

    def get_project(self):
        if (len(sys.argv) > 1):
            self.project_id = sys.argv[1]
            if self.project_id.lower() in ("sf", "smartfreight", "smart_freight"):
                self.project_id = "TRN.HMR10052"
                self.master_sheet = "Smart Freight"
            elif self.project_id.lower() in ("ott", "tariff"):
                self.project_id = "TRN.HMR11896"
                self.master_sheet = "OTT"
            elif self.project_id.lower() in ("sf_support", "support"):
                self.project_id = "TRN.HMR12175"
                self.master_sheet = "SF_SUPPORT"
            else:
                self.project_id = "TRN.HMR10142"
                self.master_sheet = "STW"
        else:
            self.project_id = "TRN.HMR10052"
            self.master_sheet = "Smart Freight"

    def get_timesheet_month(self):
        if (len(sys.argv) > 2):
            if sys.argv[2] == "all":
                self.timesheet_month = ""
            else:
                self.timesheet_month = sys.argv[2].lower()
        else:
            self.timesheet_month = ""

    def convert_excel(self):
        if (len(sys.argv) > 3):
            if (sys.argv[3] != "refresh"):
                print("Converting source Excel to CSV")
                read_file = pd.read_excel(self.src, self.sheet)
                read_file.to_csv(self.dest, index=None, header=True)
                print("Conversion complete")

    def get_engine_timesheet_data(self):
        with open(self.dest) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if line_count != 0:
                    entry = timesheet_entry()
                    entry.project_id = row[0]
                    entry.resource_name = row[16]
                    entry.resource_type = row[17]
                    entry.timesheet_period = row[24]
                    entry.day_date = row[26]
                    entry.hours = row[31]

                    if entry.project_id == self.project_id:
                        entry.fix_hours()
                        entry.get_hmrc_hours()
                        entry.process_date()
                        entry.get_day_rate(self.resources)
                        entry.get_sow(self.sows)
                        self.match_count += 1
                        if self.timesheet_month == "":
                            self.entries.append(entry)
                        else:
                            if entry.year == "2021":
                                if entry.month.lower() == self.timesheet_month:
                                    self.entries.append(entry)

                line_count += 1
        self.combine_like_lines()

    def combine_like_lines(self):
        self.entries.sort(key=lambda x: x.resource_name, reverse=False)
        self.entries.sort(key=lambda x: x.day_date, reverse=False)
        for i in range(1, len(self.entries)):
            e1 = self.entries[i-1]
            e2 = self.entries[i]
            if (e1.resource_name == e2.resource_name) and (e1.day_date == e2.day_date):
                e1.hours_fixed += e2.hours_fixed
                e2.mark_for_deletion = True

        index = -1
        for entry in self.entries:
            index += 1
            if entry.mark_for_deletion == True:
                del self.entries[index]

        self.entries.sort(key=lambda x: x.day_date, reverse=False)
        self.entries.sort(key=lambda x: x.resource_name, reverse=False)

    def write_timesheet_data(self):
        if self.match_count > 0:
            print("Writing timesheet data")
            wb = openpyxl.Workbook()

            # Write the Raw data sheet
            sheet = wb.active
            sheet.title = 'Raw'

            self.write_headers(sheet)

            index = 1
            for entry in self.entries:
                index += 1
                sheet.cell(row=index, column=1).value = entry.project_id
                sheet.cell(row=index, column=2).value = entry.resource_name
                sheet.cell(row=index, column=3).value = entry.resource_type
                sheet.cell(row=index, column=4).value = entry.hours
                sheet.cell(row=index, column=5).value = entry.hours_fixed
                sheet.cell(row=index, column=6).value = entry.hours_hmrc
                sheet.cell(row=index, column=7).value = str(entry.day_date_formatted).replace(" 00:00:00", "")
                sheet.cell(row=index, column=8).value = entry.day_of_week
                sheet.cell(row=index, column=9).value = entry.month
                sheet.cell(row=index, column=10).value = str(entry.timesheet_period_formatted).replace(" 00:00:00", "")
                sheet.cell(row=index, column=11).value = entry.day_rate
                sheet.cell(row=index, column=12).value = entry.cost()
                sheet.cell(row=index, column=13).value = entry.sow_id

                if entry.resource_type == "Employee":
                    if entry.hours_fixed not in self.permitted_employee:
                        sheet.cell(row=index, column=5).fill = PatternFill(
                            start_color="FFFF99", end_color="FFFF99", fill_type="solid")
                else:
                    if entry.hours_fixed not in self.permitted_contractor:
                        sheet.cell(row=index, column=5).fill = PatternFill(
                            start_color="FFFF99", end_color="FFFF99", fill_type="solid")

            self.format_columns(sheet)
            # self.create_pivot_sheets(wb)

            # Make the dest folder
            dest = "dest"
            if not os.path.isdir(dest):
                os.mkdir(dest)

            # Save the file
            if self.timesheet_month != "":
                self.filename = "dest/" + self.project_id + \
                    "_" + self.timesheet_month.title() + ".xlsx"
            else:
                self.filename = "dest/" + self.project_id + ".xlsx"
            wb.save(self.filename)

            # self.write_pivots()

    def format_columns(self, sheet):
        column_widths = [15, 20, 15, 15, 15, 15, 20, 15, 15, 15, 20, 20]
        index = 1
        for w in column_widths:
            sheet.column_dimensions[chr(index + 64)].width = w
            index += 1

        sheet.freeze_panes = 'B2'
        sheet.auto_filter.ref = sheet.dimensions

    def create_pivot_sheets(self, wb):
        # return
        ws_week = wb.create_sheet("By week")
        ws1_month = wb.create_sheet("By month")

    def write_headers(self, sheet):
        sheet.cell(row=1, column=1).value = "Project ID"
        sheet.cell(row=1, column=2).value = "Resource name"
        sheet.cell(row=1, column=3).value = "Resource type"
        sheet.cell(row=1, column=4).value = "Hours"
        sheet.cell(row=1, column=5).value = "Hours (fixed)"
        sheet.cell(row=1, column=6).value = "Hours (HMRC)"
        sheet.cell(row=1, column=7).value = "Day"
        sheet.cell(row=1, column=8).value = "Day of week"
        sheet.cell(row=1, column=9).value = "Month"
        sheet.cell(row=1, column=10).value = "Timesheet period (start)"
        sheet.cell(row=1, column=11).value = "Day rate"
        sheet.cell(row=1, column=12).value = "Charge amount"
        sheet.cell(row=1, column=13).value = "SOW ID"

        sheet['A1'].font = Font(bold=True)
        sheet['B1'].font = Font(bold=True)
        sheet['C1'].font = Font(bold=True)
        sheet['D1'].font = Font(bold=True)
        sheet['E1'].font = Font(bold=True)
        sheet['F1'].font = Font(bold=True)
        sheet['G1'].font = Font(bold=True)
        sheet['H1'].font = Font(bold=True)
        sheet['I1'].font = Font(bold=True)
        sheet['J1'].font = Font(bold=True)
        sheet['K1'].font = Font(bold=True)
        sheet['L1'].font = Font(bold=True)

    def get_resources(self):
        # Get the resource day rates
        print("Getting resource rates")
        with open("src/charge_rates.csv") as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if line_count == 0:
                    pass
                else:
                    if len(row) > 1:
                        res = resource()
                        res.project_id = row[0]
                        res.resource_name = row[1]
                        res.day_rate = int(row[2])
                        res.resource_type = row[3]
                        res.charge_under = row[4]
                        res.get_hours_per_day()
                        if res.charge_under == "":
                            res.charge_under = res.resource_name
                        self.resources.append(res)
                line_count += 1

    def get_sows(self):
        print("Getting statements of work")
        with open("src/sows.csv") as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if line_count == 0:
                    pass
                else:
                    my_sow = sow()
                    my_sow.project_id = row[0]
                    my_sow.sow_id = row[1]
                    my_sow.start_date = row[2]
                    my_sow.end_date = row[3]
                    my_sow.format_dates()
                    self.sows.append(my_sow)
                line_count += 1

    def write_pivots(self):
        # Pivot the data
        print("Writing pivot tables")
        df = pd.read_excel(self.filename, parse_dates=[
                           'Timesheet period (start)', 'Day'])
        df['Hours (fixed)'].round(decimals=4)
        print("\nBy period")
        pivot = pd.pivot_table(df, index='Resource name', columns='Timesheet period (start)',
                               values='Charge amount', aggfunc='sum', fill_value='0', margins=True)
        print(pivot)

        print("\n\nBy hours")
        pivot = pd.pivot_table(df, index='Resource name', columns='Timesheet period (start)',
                               values='Hours (fixed)', aggfunc='sum', fill_value='0', margins=False)
        print(pivot)

        pivot = pd.pivot_table(df, index='Resource name', columns='Month',
                               values='Charge amount', aggfunc='sum', fill_value='0', margins=True)
        print(pivot)
        return

        writer = pd.ExcelWriter(self.filename, engine='xlsxwriter')
        pivot.to_excel(writer, sheet_name='By week', header=True, index=True)
        writer.save()

    def terminate(self):
        print("Conversion complete")

    def get_hmrc_data(self):
        print("Getting HMRC timesheets for comparison")
        read_file = pd.read_excel(self.master, self.master_sheet)
        read_file.to_csv(self.master_dest, index=None, header=True)
        with open(self.master_dest) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if line_count == 0:
                    self.get_master_top_row(row)
                else:
                    self.get_master_row(row)

                line_count += 1

    def get_master_top_row(self, row):
        my_resource = None
        for res in self.resources:
            matched = False
            col_count = 0
            for item in row:
                col_count += 1
                if col_count >= 6:
                    if res.charge_under == item:
                        if res.project_id == self.project_id:
                            matched = True
                            my_resource = res
                            break

            if matched == True:
                my_resource.match_column = col_count

    def get_master_row(self, row):
        if (row[0] != ""):
            timesheet_date = row[0]
            timesheet_period = row[4]
            today = date.today()
            if datetime.strptime(timesheet_date, '%Y-%m-%d').date() <= today:
                col_count = 0
                for item in row:
                    col_count += 1
                    if col_count >= 6:
                        entry = hmrc_timesheet_entry()
                        entry.project_id = self.project_id
                        entry.day_date = timesheet_date
                        entry.timesheet_period = timesheet_period
                        entry.get_resource_name(self, col_count)
                        entry.process_date()
                        entry.hours = item
                        entry.fix_hours()
                        if entry.resource_name != "":
                            self.hmrc_entries.append(entry)

    def timesheets_match(self, a, b):
        matched = False
        if a.day_date_as_date == b.day_date_as_date:
            if a.project_id == b.project_id:
                if a.charge_under == b.resource_name:
                    matched = True
        return (matched)

    def compare_timesheets(self):
        # This looks for timesheets in the Engine list and compares against HMRC timesheets
        # looks for missing timesheets and mismatched of numbers
        with open('src/tmp_compare.csv', 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Timesheet date", "Resource name",
                             "Engine hours", "HMRC hours"])

            for entry in self.entries:
                matched = False
                for hmrc_entry in self.hmrc_entries:
                    if self.timesheets_match(entry, hmrc_entry):
                        # print("found match")
                        matched = True
                        break

                if matched == False:
                    print("No match on timesheet by " +
                          entry.resource_name + " on date " + entry.day_date)
                elif matched == True:
                    hrs_per_day = hmrc_entry.resource.hours_per_day
                    if hmrc_entry.resource.hours_per_day == 7.5:
                        hmrc_entry.hours2 = hmrc_entry.hours * 7.5 / 8
                    else:
                        hmrc_entry.hours2 = hmrc_entry.hours
                    if hmrc_entry.hours2 != entry.hours_fixed:
                        writer.writerow([entry.day_date, entry.resource_name, entry.hours_fixed,
                                         hmrc_entry.resource.resource_name, hmrc_entry.hours])
